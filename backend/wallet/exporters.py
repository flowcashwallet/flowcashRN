import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.http import HttpResponse
from datetime import datetime
from io import BytesIO

def export_transactions_to_excel(transactions):
    """
    Generates an Excel file from a queryset of transactions with a split view (Income vs Expenses).
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Transactions"

    # Separate data
    incomes = [t for t in transactions if t.type == 'income']
    expenses = [t for t in transactions if t.type == 'expense']

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    sub_header_font = Font(bold=True, color="FFFFFF")
    
    green_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid") # Dark Green
    red_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid") # Dark Red
    
    # Main Headers
    worksheet.merge_cells('A1:C1')
    cell_a1 = worksheet['A1']
    cell_a1.value = "INGRESOS (INCOME)"
    cell_a1.font = header_font
    cell_a1.fill = green_fill
    cell_a1.alignment = Alignment(horizontal='center')

    worksheet.merge_cells('E1:G1')
    cell_e1 = worksheet['E1']
    cell_e1.value = "GASTOS (EXPENSES)"
    cell_e1.font = header_font
    cell_e1.fill = red_fill
    cell_e1.alignment = Alignment(horizontal='center')

    # Sub Headers
    columns = [
        ('A', "Fecha", green_fill), ('B', "Descripción", green_fill), ('C', "Monto", green_fill),
        ('E', "Fecha", red_fill), ('F', "Descripción", red_fill), ('G', "Monto", red_fill)
    ]
    
    for col, title, fill in columns:
        cell = worksheet[f'{col}2']
        cell.value = title
        cell.font = sub_header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')

    # Data
    max_rows = max(len(incomes), len(expenses))
    
    total_income = 0
    total_expense = 0
    
    border_bottom = Border(bottom=Side(style='thin', color="CCCCCC"))

    for i in range(max_rows):
        row_num = i + 3
        
        # Income
        if i < len(incomes):
            inc = incomes[i]
            total_income += inc.amount
            
            # Date
            cell_date = worksheet[f'A{row_num}']
            cell_date.value = inc.date.strftime('%Y-%m-%d') if inc.date else ""
            cell_date.border = border_bottom
            cell_date.alignment = Alignment(horizontal='center')

            # Description (combining Category + Description for context)
            desc = f"{inc.category or 'General'}: {inc.description}"
            cell_desc = worksheet[f'B{row_num}']
            cell_desc.value = desc
            cell_desc.border = border_bottom
            
            # Amount
            cell_amt = worksheet[f'C{row_num}']
            cell_amt.value = inc.amount
            cell_amt.number_format = '#,##0.00'
            cell_amt.font = Font(color="2E7D32") # Green text
            cell_amt.border = border_bottom
        
        # Expense
        if i < len(expenses):
            exp = expenses[i]
            total_expense += exp.amount
            
            # Date
            cell_date = worksheet[f'E{row_num}']
            cell_date.value = exp.date.strftime('%Y-%m-%d') if exp.date else ""
            cell_date.border = border_bottom
            cell_date.alignment = Alignment(horizontal='center')

            # Description
            desc = f"{exp.category or 'General'}: {exp.description}"
            cell_desc = worksheet[f'F{row_num}']
            cell_desc.value = desc
            cell_desc.border = border_bottom
            
            # Amount
            cell_amt = worksheet[f'G{row_num}']
            cell_amt.value = exp.amount
            cell_amt.number_format = '#,##0.00'
            cell_amt.font = Font(color="C62828") # Red text
            cell_amt.border = border_bottom

    # Totals
    total_row = max_rows + 4
    
    thick_top = Border(top=Side(style='medium'))

    # Income Total
    worksheet[f'A{total_row}'] = "TOTAL INGRESOS"
    worksheet[f'A{total_row}'].font = Font(bold=True)
    worksheet[f'A{total_row}'].border = thick_top
    
    cell = worksheet[f'C{total_row}']
    cell.value = total_income
    cell.number_format = '$#,##0.00'
    cell.font = Font(bold=True, color="2E7D32")
    cell.border = thick_top

    # Expense Total
    worksheet[f'E{total_row}'] = "TOTAL GASTOS"
    worksheet[f'E{total_row}'].font = Font(bold=True)
    worksheet[f'E{total_row}'].border = thick_top
    
    cell = worksheet[f'G{total_row}']
    cell.value = total_expense
    cell.number_format = '$#,##0.00'
    cell.font = Font(bold=True, color="C62828")
    cell.border = thick_top

    # Net Balance
    net_row = total_row + 2
    net_val = total_income - total_expense
    worksheet.merge_cells(f'A{net_row}:G{net_row}')
    cell = worksheet[f'A{net_row}']
    cell.value = f"BALANCE NETO: ${net_val:,.2f}"
    cell.font = Font(bold=True, size=14, color="000000" if net_val >= 0 else "C62828")
    cell.alignment = Alignment(horizontal='center')

    # Column Widths
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 3 # Spacer
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 30
    worksheet.column_dimensions['G'].width = 15

    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=transactions_split_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    workbook.save(response)
    return response

def export_vision_to_excel(entities):
    """
    Generates an Excel file from a queryset of VisionEntities (Assets/Liabilities) in split view.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Balance Sheet"

    # Separate data
    assets = [e for e in entities if e.type == 'asset']
    liabilities = [e for e in entities if e.type == 'liability']
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    sub_header_font = Font(bold=True, color="FFFFFF")
    
    green_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid") # Dark Green
    red_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid") # Dark Red
    
    # Main Headers
    worksheet.merge_cells('A1:B1')
    cell_a1 = worksheet['A1']
    cell_a1.value = "ACTIVOS (ASSETS)"
    cell_a1.font = header_font
    cell_a1.fill = green_fill
    cell_a1.alignment = Alignment(horizontal='center')

    worksheet.merge_cells('D1:E1')
    cell_d1 = worksheet['D1']
    cell_d1.value = "PASIVOS (LIABILITIES)"
    cell_d1.font = header_font
    cell_d1.fill = red_fill
    cell_d1.alignment = Alignment(horizontal='center')

    # Sub Headers
    columns = [
        ('A', "Nombre", green_fill), ('B', "Monto", green_fill),
        ('D', "Nombre", red_fill), ('E', "Monto", red_fill)
    ]
    
    for col, title, fill in columns:
        cell = worksheet[f'{col}2']
        cell.value = title
        cell.font = sub_header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')

    # Data
    max_rows = max(len(assets), len(liabilities))
    
    total_assets = 0
    total_liabilities = 0
    
    border_bottom = Border(bottom=Side(style='thin', color="CCCCCC"))

    for i in range(max_rows):
        row_num = i + 3
        
        # Asset
        if i < len(assets):
            asset = assets[i]
            total_assets += asset.amount
            
            cell_name = worksheet[f'A{row_num}']
            cell_name.value = asset.name
            cell_name.border = border_bottom
            
            cell_amt = worksheet[f'B{row_num}']
            cell_amt.value = asset.amount
            cell_amt.number_format = '#,##0.00'
            cell_amt.font = Font(color="2E7D32") # Green text
            cell_amt.border = border_bottom
        
        # Liability
        if i < len(liabilities):
            liab = liabilities[i]
            total_liabilities += liab.amount
            
            cell_name = worksheet[f'D{row_num}']
            cell_name.value = liab.name
            cell_name.border = border_bottom
            
            cell_amt = worksheet[f'E{row_num}']
            cell_amt.value = liab.amount
            cell_amt.number_format = '#,##0.00'
            cell_amt.font = Font(color="C62828") # Red text
            cell_amt.border = border_bottom

    # Totals
    total_row = max_rows + 4
    thick_top = Border(top=Side(style='medium'))
    
    # Asset Total
    worksheet[f'A{total_row}'] = "TOTAL ACTIVOS"
    worksheet[f'A{total_row}'].font = Font(bold=True)
    worksheet[f'A{total_row}'].border = thick_top
    
    cell = worksheet[f'B{total_row}']
    cell.value = total_assets
    cell.number_format = '$#,##0.00'
    cell.font = Font(bold=True, color="2E7D32")
    cell.border = thick_top

    # Liability Total
    worksheet[f'D{total_row}'] = "TOTAL PASIVOS"
    worksheet[f'D{total_row}'].font = Font(bold=True)
    worksheet[f'D{total_row}'].border = thick_top
    
    cell = worksheet[f'E{total_row}']
    cell.value = total_liabilities
    cell.number_format = '$#,##0.00'
    cell.font = Font(bold=True, color="C62828")
    cell.border = thick_top

    # Net Worth
    net_row = total_row + 2
    worksheet.merge_cells(f'A{net_row}:E{net_row}')
    cell = worksheet[f'A{net_row}']
    cell.value = f"PATRIMONIO NETO: ${total_assets - total_liabilities:,.2f}"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')

    # Column Widths
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 3 # Spacer
    worksheet.column_dimensions['D'].width = 30
    worksheet.column_dimensions['E'].width = 15

    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=balance_sheet_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    workbook.save(response)
    return response

def export_transactions_to_pdf(transactions):
    """
    Generates a PDF file from a queryset of transactions.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = styles['Title']
    elements.append(Paragraph(f"Transaction Report - {datetime.now().strftime('%Y-%m-%d')}", title_style))
    elements.append(Spacer(1, 12))

    # Table Data
    data = [["Date", "Description", "Category", "Type", "Payment", "Amount"]]
    
    total_income = 0
    total_expense = 0

    for t in transactions:
        date_str = t.date.strftime("%Y-%m-%d %H:%M") if t.date else ""
        amount = float(t.amount)
        
        if t.type == 'income':
            total_income += amount
        elif t.type == 'expense':
            total_expense += amount
            
        row = [
            date_str,
            t.description[:30] + ('...' if len(t.description) > 30 else ''), # Truncate long descriptions
            t.category or "-",
            t.get_type_display(),
            t.get_payment_type_display() if t.payment_type else "-",
            f"${amount:,.2f}"
        ]
        data.append(row)

    # Table Style
    table = Table(data, colWidths=[110, 200, 100, 80, 100, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'), # Left align description
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'), # Right align amount
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 24))

    # Summary
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=6
    )
    
    elements.append(Paragraph("<b>Summary:</b>", summary_style))
    elements.append(Paragraph(f"Total Income: ${total_income:,.2f}", summary_style))
    elements.append(Paragraph(f"Total Expense: ${total_expense:,.2f}", summary_style))
    elements.append(Paragraph(f"Net Balance: ${total_income - total_expense:,.2f}", summary_style))

    # Build PDF
    doc.build(elements)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=transactions_{datetime.now().strftime("%Y%m%d")}.pdf'
    response.write(pdf)
    return response

def export_vision_to_pdf(entities):
    """
    Generates a PDF file from a queryset of VisionEntities.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = styles['Title']
    elements.append(Paragraph(f"Balance Sheet - {datetime.now().strftime('%Y-%m-%d')}", title_style))
    elements.append(Spacer(1, 12))

    # Table Data
    data = [["Name", "Type", "Category", "Amount"]]
    
    total_assets = 0
    total_liabilities = 0

    for entity in entities:
        type_display = entity.get_type_display()
        if entity.type == 'asset':
            total_assets += entity.amount
        elif entity.type == 'liability':
            total_liabilities += entity.amount
            
        data.append([
            entity.name,
            type_display,
            entity.category or "",
            f"{entity.amount:,.2f}"
        ])

    # Create Table
    table = Table(data, colWidths=[150, 100, 100, 100])
    
    # Table Style
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)
    
    elements.append(table)
    elements.append(Spacer(1, 24))

    # Summary
    summary_style = styles['Normal']
    net_worth = total_assets - total_liabilities
    
    elements.append(Paragraph(f"<b>Total Assets:</b> {total_assets:,.2f}", summary_style))
    elements.append(Paragraph(f"<b>Total Liabilities:</b> {total_liabilities:,.2f}", summary_style))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"<b>Net Worth:</b> {net_worth:,.2f}", summary_style))

    # Build PDF
    doc.build(elements)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=balance_sheet_{datetime.now().strftime("%Y%m%d")}.pdf'
    response.write(pdf)
    return response
